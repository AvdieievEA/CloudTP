import os
from pprint import pprint
from datetime import timedelta, datetime
import telebot
import config
import urllib.request

from openpyxl import load_workbook
import pandas as pd

bot = telebot.TeleBot(config.token)


@bot.message_handler(commands=['start'])
def hi(message):
    bot.send_message(message.chat.id, 'Привет, скинь мне файл, который нужно обработать')


@bot.message_handler(content_types=['document'])
def handle_docs(message):
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        src = 'C:/Users/avdie/Desktop/Project/ModelTP/' + message.document.file_name;
        with open(src, 'wb') as new_file:
            new_file.write(downloaded_file)

        bot.reply_to(message, "Пожалуйста, дайте мне минутку")
        os.chdir(r"C:\Users\avdie\Desktop\Project\ModelTP")

        FILE = message.document.file_name

        WORKBOOK = load_workbook(FILE)
        ko_read = WORKBOOK['KO']
        zo_read = WORKBOOK['ZO']
        period_read = WORKBOOK['Period']

        def get_periods_timedelta():
            WORKBOOK.active = 0
            periods = [period_read.cell(row=row, column=2).value for row in range(2, period_read.max_row + 1)]
            return [
                {"interval": timedelta(days=period), "interval_type": period_read.cell(row=i + 2, column=3).value}
                for i, period in enumerate(periods)
            ]

        def check_workday(date, period):
            date_counter = timedelta(days=0)
            while date_counter != periods[index_period]["interval"]:
                date += period
                if date.weekday() < 5:
                    date_counter += timedelta(days=1)
            return date

        periods = get_periods_timedelta()
        WORKBOOK.active = 1
        nested_dicts_ko = []
        index_period = 0
        for i in range(0, len(periods)):
            nested_dict_ko = []
            nested_dicts_ko.append(nested_dict_ko)
            for row in range(2, ko_read.max_row + 1):
                data_postavka_ko = ko_read.cell(row=row, column=5).value
                x = {
                    "napravlenie": ko_read.cell(row=row, column=2).value,
                    "data": data_postavka_ko,
                    "interval": periods[index_period],
                    "product": ko_read.cell(row=row, column=3).value,
                    "weight": ko_read.cell(row=row, column=7).value,
                    "price_val": ko_read.cell(row=row, column=6).value,
                    "price_uah": ko_read.cell(row=row, column=8).value,
                    "fare": ko_read.cell(row=row, column=9).value,
                    "discont": ko_read.cell(row=row, column=10).value,
                    "index_row": row - 1,
                    "nested_dict_index": i,
                    "zistavni": [],
                    "valuta": ko_read.cell(row=row, column=11).value,
                    "suma_valuta": ko_read.cell(row=row, column=12).value,
                    "kurs": ko_read.cell(row=row, column=8).value / ko_read.cell(row=row, column=12).value,
                }
                if periods[index_period]["interval_type"] == "Календарний день":
                    x.update({
                        "data_min": data_postavka_ko - periods[index_period]["interval"],
                        "data_max": data_postavka_ko + periods[index_period]["interval"],
                    })
                elif periods[index_period]["interval_type"] == "Робочий день":
                    x["data_min"] = check_workday(data_postavka_ko, -timedelta(days=1))
                    x["data_max"] = check_workday(data_postavka_ko, timedelta(days=1))
                nested_dict_ko.append(x)
            index_period += 1
        list_zo = [
            {
                "product": zo_read.cell(row=row, column=2).value,
                "data": zo_read.cell(row=row, column=4).value,
                "price": zo_read.cell(row=row, column=5).value,
                "kontragent": zo_read.cell(row=row, column=6).value,
                "index": zo_read.cell(row=row, column=1).value,
                "zistavnist": zo_read.cell(row=row, column=7).value
            }
            for row in range(2, zo_read.max_row + 1)
        ]
        for nested_dict_ko in nested_dicts_ko:
            for ko in nested_dict_ko:
                for zo in list_zo:
                    if zo["product"] == ko["product"] and ko["data_min"] <= zo["data"] <= ko["data_max"] and zo[
                        "zistavnist"] == "Зіставна операція":
                        ko["zistavni"].append(zo)

        for nested_dict_ko in nested_dicts_ko:
            for ko in nested_dict_ko:
                for zo in ko["zistavni"]:
                    while True:
                        swap_counter = 0
                        for i in range(len(ko["zistavni"]) - 1):
                            if ko["zistavni"][i]["price"] > ko["zistavni"][i + 1]["price"]:
                                a = ko["zistavni"][i]
                                ko["zistavni"][i] = ko["zistavni"][i + 1]
                                ko["zistavni"][i + 1] = a
                                swap_counter += 1
                        if swap_counter == 0:
                            break

        for nested_dict_ko in nested_dicts_ko:
            for ko in nested_dict_ko:
                if len(ko["zistavni"]) == 0:
                    ko["diapazon_min"] = "NA"
                    ko["diapazon_max"] = "NA"
                elif len(ko["zistavni"]) % 4 == 0:
                    ko["diapazon_min"] = (ko["zistavni"][int(len(ko["zistavni"]) // 4 - 1)]["price"] +
                                          ko["zistavni"][int(len(ko["zistavni"]) // 4)]["price"]) / 2
                    ko["diapazon_max"] = (ko["zistavni"][int(len(ko["zistavni"]) * 0.75)]["price"] +
                                          ko["zistavni"][int(len(ko["zistavni"]) * 0.75 - 1)]["price"]) / 2
                else:
                    ko["diapazon_min"] = ko["zistavni"][int(len(ko["zistavni"]) // 4)]["price"]
                    ko["diapazon_max"] = ko["zistavni"][int((len(ko["zistavni"]) * 0.75) // 1)]["price"]

        testp = 1
        for nested_dict_ko in nested_dicts_ko:
            for ko in nested_dict_ko:
                if ko["diapazon_min"] == 'NA':
                    ko['status'] = "Немає зіставних"
                    ko['donarahuvannya_baza'] = 0
                    ko['donarahuvannya'] = 0
                else:
                    if ko['napravlenie'] == "Експорт":
                        diapazon_min_uah = ((ko["diapazon_min"] + ko["fare"]) * ko["weight"] * ko['kurs']) * ko[
                            'discont']
                        diapazon_max_uah = ((ko["diapazon_max"] + ko["fare"]) * ko["weight"] * ko['kurs']) * ko[
                            'discont']
                        if ko['price_uah'] < diapazon_min_uah:
                            ko['status'] = "Нижче ринкового діапазону"
                            ko['donarahuvannya_baza'] = diapazon_min_uah - ko['price_uah']
                            ko['donarahuvannya'] = (diapazon_min_uah - ko['price_uah']) * 0.18
                        elif diapazon_min_uah < ko['price_uah'] < diapazon_max_uah:
                            ko['status'] = "В межах ринкового діапазону"
                            ko['donarahuvannya_baza'] = 0
                            ko['donarahuvannya'] = 0
                        elif ko['price_uah'] > diapazon_max_uah:
                            ko['status'] = "Вище ринкового діапазону"
                            ko['donarahuvannya_baza'] = 0
                            ko['donarahuvannya'] = 0
                    if ko['napravlenie'] == "Імпорт":
                        diapazon_min_uah = ((ko["diapazon_min"] - ko["fare"]) * ko["weight"] * ko['kurs']) * ko[
                            'discont']
                        diapazon_max_uah = ((ko["diapazon_max"] - ko["fare"]) * ko["weight"] * ko['kurs']) * ko[
                            'discont']
                        if ko['price_uah'] < diapazon_min_uah:
                            ko['status'] = "Нижче ринкового діапазону"
                            ko['donarahuvannya_baza'] = 0
                            ko['donarahuvannya'] = 0
                        elif diapazon_min_uah < ko['price_uah'] < diapazon_max_uah:
                            ko['status'] = "В межах ринкового діапазону"
                            ko['donarahuvannya_baza'] = 0
                            ko['donarahuvannya'] = 0
                        elif ko['price_uah'] > diapazon_max_uah:
                            ko['status'] = "Вище ринкового діапазону"
                            ko['donarahuvannya_baza'] = ko['price_uah'] - diapazon_max_uah
                            ko['donarahuvannya'] = (ko['price_uah'] - diapazon_max_uah) * 0.18
                ko['donarahuvannya'] = float(ko['donarahuvannya'])

        # TODO comprehension
        don_sum = {}

        for nested_dict_ko in nested_dicts_ko:
            suma = 0
            na_counter = 0
            for ko in nested_dict_ko:
                suma += ko["donarahuvannya"]
                interval = str(ko['interval']['interval'].days)
                if ko['status'] == "Немає зіставних":
                    na_counter += 1
            don_sum[interval] = {
                'NA': na_counter,
                'donarahuvannya': float(suma)
            }

        ko_pd = {
            "Номер поставки": [],
            "Інтервал": [],
            "Тип інтервалу": [],
            "Номенклатура": [],
            "Дата поставки": [],
            "Ціна товару, валюта/т": [],
            "Обсяг поставки, т": [],
            "Вартість поставки (факт), грн., FCA": [],
            "Дата зіставних операцій, від": [],
            "Дата зіставних операцій, до": [],
            "Ринковий діапазон, від": [],
            "Ринковий діапазон, до": [],
            "Статус": [],
            "Різниця вартості поставки з ринковим діапазоном, грн": [],
            "Донарахування": [],
            "Кількість зіставних операцій": []
        }

        zo_pd = {
            "№ поставки КО": [],
            "Інтервал": [],
            "Тип інтервалу": [],
            "Контрагент": [],
            "Номенклатура": [],
            'Дата поставки': [],
            'Ціна товару, валюта': [],
            'Зіставність': []
        }

        titul_pd = {
            "Інтервал": [],
            "Доначисления": [],
            "Количество поставок без сопоставимых": []
        }

        for i in don_sum.keys():
            titul_pd['Інтервал'].append(i)

        for i in don_sum.keys():
            titul_pd["Доначисления"].append(don_sum[i]['donarahuvannya'])
            titul_pd["Количество поставок без сопоставимых"].append(don_sum[i]['NA'])

        pprint(don_sum)

        for nested_dict_ko in nested_dicts_ko:
            for ko in nested_dict_ko:
                ko_pd["Номер поставки"].append(ko['index_row'])
                ko_pd["Інтервал"].append(ko['interval']['interval'].days)
                ko_pd["Тип інтервалу"].append(ko['interval']['interval_type'])
                ko_pd["Номенклатура"].append(ko['product'])
                ko_pd["Дата поставки"].append(ko['data'])
                ko_pd["Ціна товару, валюта/т"].append(ko['price_val'])
                ko_pd["Обсяг поставки, т"].append(ko['weight'])
                ko_pd["Вартість поставки (факт), грн., FCA"].append(ko['price_uah'])
                ko_pd["Дата зіставних операцій, від"].append(ko['data_min'])
                ko_pd["Дата зіставних операцій, до"].append(ko['data_max'])
                ko_pd["Ринковий діапазон, від"].append(ko['diapazon_min'])
                ko_pd["Ринковий діапазон, до"].append(ko['diapazon_max'])
                ko_pd["Статус"].append(ko['status'])
                ko_pd["Різниця вартості поставки з ринковим діапазоном, грн"].append(ko['donarahuvannya_baza'])
                ko_pd["Донарахування"].append(ko['donarahuvannya'])
                ko_pd["Кількість зіставних операцій"].append(len(ko['zistavni']))
                for z in ko['zistavni']:
                    zo_pd["№ поставки КО"].append(ko['index_row'])
                    zo_pd["Інтервал"].append(ko['interval']['interval'].days)
                    zo_pd["Тип інтервалу"].append(ko['interval']['interval_type'])
                    zo_pd["Контрагент"].append(z['kontragent'])
                    zo_pd["Номенклатура"].append(z['product'])
                    zo_pd['Дата поставки'].append(z['data'])
                    zo_pd['Ціна товару, валюта'].append(z['price'])
                    zo_pd['Зіставність'].append(z['zistavnist'])

        result_sheets = {
            "Короткий опис": pd.DataFrame(titul_pd),
            "Контрольовані операції": pd.DataFrame(ko_pd),
            "Зіставні операції": pd.DataFrame(zo_pd)
        }

        pprint(nested_dicts_ko)

        writer = pd.ExcelWriter(r"C:\Users\avdie\Desktop\Project\ModelTP\Шаблон для моделі підбору цін_DMD_2_итог.xlsx",
                                engine='xlsxwriter')
        for sheet_name in result_sheets.keys():
            result_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

        writer.save()
        file_final = open(rf"C:\Users\avdie\Desktop\Project\ModelTP\{message.document.file_name}_итог.xlsx", 'rb')
        bot.send_document(message.chat.id, file_final)
    except Exception:
        bot.reply_to(message, "Пожалуйста, попробуйте удалить 300-400 пустых строк на вкладке КО, имеено удаляя строки, а не очищия их")
        file_help = open(rf"C:\Users\avdie\Desktop\Project\ModelTP\Інструкція.png", 'rb')
        bot.send_photo(message.chat.id, file_help)

bot.polling()
